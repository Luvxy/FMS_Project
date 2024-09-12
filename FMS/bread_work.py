import pandas as pd
from itertools import combinations
from openpyxl import load_workbook
import hwp_test as hp

# 엑셀 파일 경로
file_path = 'C:/Users/user/Documents/GitHub/FMS_Project/FMS/bread_split.xlsx'  # 업로드한 엑셀 파일 경로

# 엑셀 파일에서 데이터 읽기
df = pd.read_excel(file_path, sheet_name='Sheet1')

# 1열, 2열, 3열, 4열의 데이터를 가져오기
left_names = df.iloc[:, 0].dropna().tolist()  # 1열 데이터
left_numbers = df.iloc[:, 1].dropna().tolist()  # 2열 데이터 (NaN 제거)
right_names = df.iloc[:, 2].dropna().tolist()  # 3열 데이터
right_numbers = df.iloc[:, 3].dropna().tolist()  # 4열 데이터 (NaN 제거)

def distribute_numbers(left_numbers, right_numbers):
    result = []
    original_right_numbers = right_numbers.copy()  # 원본을 유지하기 위해 복사본 사용
    unallocated_left = []  # 분배되지 못한 left 값 저장
    unallocated_right = right_numbers.copy()  # 분배되지 못한 right 값 저장

    for left in left_numbers:
        found = False
        best_fit = None
        re_comb = []
        re_left = left
        
        for i in range(1, len(right_numbers) + 1):
            for comb in combinations(right_numbers, i):
                comb_sum = sum(comb)
                if comb_sum == left:
                    result.append((left, list(comb)))
                    for num in comb:
                        right_numbers.remove(num)
                    found = True
                    break
                elif comb_sum > left:
                    if best_fit is None or comb_sum < sum(best_fit):
                        best_fit = comb
            if found:
                break

        if not found:
            if best_fit:
                result.append((left, list(best_fit)))
                for num in best_fit:
                    right_numbers.remove(num)
            else:
                # 분배되지 못한 left와 대응할 수 없는 right를 추적
                unallocated_left.append(left)

    # 마지막에 분배되지 못한 left와 remaining right를 한 번에 묶어줌
    if unallocated_left and right_numbers:
        result.append((unallocated_left, right_numbers))
        right_numbers.clear()

    return result, original_right_numbers


# Distribute the numbers
distributed_numbers, original_right_numbers = distribute_numbers(left_numbers, right_numbers)

# 결과를 데이터프레임으로 변환
result_data = []
used_right_numbers = []

max_comb_length = 0  # 가장 긴 조합의 길이를 추적합니다

for i, (left, right) in enumerate(distributed_numbers):
    left_name = left_names[i]
    right_names_combined = []
    right_values_combined = []
    for right_value in right:
        right_index = original_right_numbers.index(right_value)
        right_name = right_names[right_index]
        right_names_combined.append(right_name)
        right_names_combined.append(right_value)
        used_right_numbers.append(right_value)
    max_comb_length = max(max_comb_length, len(right))  # 조합의 최대 길이를 갱신합니다
    result_data.append([left_name, left] + right_names_combined)

# 결과 데이터를 데이터프레임으로 변환
columns = ['', '수량']
for i in range(max_comb_length):
    columns.append('')
for i in range(max_comb_length):
    columns.append('')

result_df = pd.DataFrame(result_data, columns=columns)

# 인수증 제작
basic_pairs = [
    ["bread1",""], ["bread2",""], ["bread3",""], ["bread4",""],
    ["bread5",""], ["bread6",""], ["bread7",""], ["bread8",""],
    ["bread9",""], ["breada",""], ["breadb",""], ["breadc",""],
    ["breadd",""], 
    ["left1",""], ["left2",""], ["left3",""], ["left4",""],
    ["left5",""], ["left6",""], ["left7",""], ["left8",""],
    ["left9",""], ["lefta",""], ["leftb",""], ["leftc",""],
    ["leftd",""], 
    ["right1",""], ["right2",""], ["right3",""], ["right4",""],
    ["right5",""], ["right6",""], ["right7",""], ["right8",""],
    ["right9",""], ["righta",""], ["rightb",""], ["rightc",""],
    ["rightd",""], 
    ["num1",""], ["num2",""], ["num3",""], ["num4",""],
    ["num5",""], ["num6",""], ["num7",""], ["num8",""],
    ["num9",""], ["numa",""], ["numb",""], ["numc",""],
    ["numd",""],
    ["date",""], ["month",""], ["day",""], ["td",""]
]

text = ""
print(result_data)

for c in range(len(result_data)):
    for i in range(len(result_data[c])):
        try:
            df = result_data[i]
            for j in range(len(df[i])-2):
                tem = df[j+2]
                tem2 = basic_pairs[i]
                if isinstance(tem, str) and j > 0:
                    text += ", " + str(tem)
                    print(text)
                elif isinstance(tem, str):
                    text += str(tem)
                    print(text)
                else:
                    text += "(" + str(tem)+")"
                    print(text)
        except:
            break
        
    a = basic_pairs[c]
    a[1] = text
    # print(basic_pairs)
    text = ""

# 기존 엑셀 파일에 새로운 시트를 추가하기 전에 기존 시트 삭제
book = load_workbook(file_path)
if 'Distribution Result' in book.sheetnames:
    del book['Distribution Result']
    book.save(file_path)

# 새로운 시트로 결과 저장
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    result_df.to_excel(writer, sheet_name='Distribution Result', index=False)

print("결과가 새로운 시트에 저장되었습니다.")

hp.creat_print_file(basic_pairs)